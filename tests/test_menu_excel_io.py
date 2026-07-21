"""Unit test menu_excel_io + MenuStore.bulk_upsert.

Jalankan: python3 -m unittest tests.test_menu_excel_io
"""

import os
import tempfile
import unittest

from app.menu_excel_io import menus_to_workbook_bytes, workbook_bytes_to_menus
from app.menu_store import MenuStore


class TestMenuExcelIO(unittest.TestCase):

    def setUp(self):
        fd, self.db_path = tempfile.mkstemp(suffix=".db")
        os.close(fd)
        self.store = MenuStore(db_path=self.db_path)
        # bersihkan seed dari _auto_migrate_from_excel (kalau ada)
        with self.store._connect() as conn:
            conn.execute("DELETE FROM menus")
            conn.commit()

    def tearDown(self):
        os.remove(self.db_path)

    # ---------- roundtrip: export -> import balik jadi menu yang sama ----------
    def test_roundtrip_export_import(self):
        rows = [
            {"name": "NASI GORENG", "price": 15000, "category": "NASI",
             "resto": "MADAM LILY", "alias": "NG", "note": "", "emoji": "🍚"},
            {"name": "TEH MANIS", "price": 5000, "category": "MINUMAN",
             "resto": "MADAM LILY", "alias": "", "note": "", "emoji": "🥤"},
        ]
        self.store.bulk_upsert(rows)

        # Export
        data = menus_to_workbook_bytes(self.store.get_all_menus())
        self.assertGreater(len(data), 100)

        # Parse balik
        parsed = workbook_bytes_to_menus(data)
        self.assertEqual(len(parsed), 2)
        by_name = {r["name"]: r for r in parsed}
        self.assertEqual(by_name["NASI GORENG"]["price"], 15000)
        self.assertEqual(by_name["TEH MANIS"]["emoji"], "🥤")

    # ---------- bulk_upsert: insert baru ----------
    def test_upsert_menu_baru(self):
        added, updated, errors = self.store.bulk_upsert([
            {"name": "SATE AYAM", "price": 20000},
        ])
        self.assertEqual((added, updated, len(errors)), (1, 0, 0))
        self.assertEqual(self.store.find_menu("SATE AYAM")["price"], 20000)

    # ---------- bulk_upsert: update yang sudah ada ----------
    def test_upsert_menu_lama(self):
        self.store.bulk_upsert([{"name": "SATE AYAM", "price": 20000}])
        added, updated, errors = self.store.bulk_upsert([
            {"name": "SATE AYAM", "price": 25000, "note": "extra pedas"},
        ])
        self.assertEqual((added, updated, len(errors)), (0, 1, 0))
        m = self.store.find_menu("SATE AYAM")
        self.assertEqual(m["price"], 25000)
        self.assertEqual(m["note"], "extra pedas")

    # ---------- bulk_upsert: case-insensitive match ----------
    def test_upsert_case_insensitive(self):
        self.store.bulk_upsert([{"name": "Sate Ayam", "price": 20000}])
        added, updated, _ = self.store.bulk_upsert([
            {"name": "SATE AYAM", "price": 22000},
        ])
        self.assertEqual((added, updated), (0, 1))

    # ---------- bulk_upsert: validasi baris tanpa nama ----------
    def test_upsert_baris_tanpa_nama_ditolak(self):
        added, updated, errors = self.store.bulk_upsert([
            {"name": "", "price": 10000},
            {"name": "TAHU", "price": 5000},
        ])
        self.assertEqual(added, 1)
        self.assertEqual(len(errors), 1)

    # ---------- bulk_upsert: validasi harga bukan angka ----------
    def test_upsert_harga_bukan_angka(self):
        added, updated, errors = self.store.bulk_upsert([
            {"name": "TAHU", "price": "gratis"},
        ])
        self.assertEqual((added, updated), (0, 0))
        self.assertEqual(len(errors), 1)

    # ---------- parse: header case-insensitive & aliases ----------
    def test_parse_header_case_insensitive(self):
        # Bikin workbook dengan header lowercase "nama"/"harga"
        import openpyxl
        import io

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["no", "resto", "kategori", "nama", "harga", "alias",
                   "catatan", "emoji"])
        ws.append([1, "MADAM LILY", "MINUMAN", "ES TEH", 5000, "", "", ""])
        buf = io.BytesIO()
        wb.save(buf)

        rows = workbook_bytes_to_menus(buf.getvalue())
        self.assertEqual(rows[0]["name"], "ES TEH")
        self.assertEqual(rows[0]["price"], 5000)

    # ---------- parse: file tanpa kolom nama ditolak ----------
    def test_parse_tanpa_kolom_nama_raise(self):
        import openpyxl
        import io

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["harga", "alias"])
        ws.append([5000, "X"])
        buf = io.BytesIO()
        wb.save(buf)

        with self.assertRaises(ValueError):
            workbook_bytes_to_menus(buf.getvalue())

    # ---------- parse: baris kosong di-skip ----------
    def test_parse_skip_baris_kosong(self):
        import openpyxl
        import io

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Nama", "Harga"])
        ws.append(["NASI", 10000])
        ws.append([None, None])
        ws.append(["", ""])
        ws.append(["TEH", 5000])
        buf = io.BytesIO()
        wb.save(buf)

        rows = workbook_bytes_to_menus(buf.getvalue())
        self.assertEqual(len(rows), 2)


if __name__ == "__main__":
    unittest.main()

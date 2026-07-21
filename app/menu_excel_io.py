"""
menu_excel_io.py

Baca & tulis daftar menu dalam format Excel (data/menu.xlsx) yang sama
dengan skema lama supaya file hasil /exportmenu bisa langsung di-import
balik lewat /importmenu.

Format kolom (baris 1 = header):
    No | Resto | Kategori | Nama | Harga | Alias | Catatan | Emoji
"""

from typing import Any, Dict, List


HEADERS = ["No", "Resto", "Kategori", "Nama", "Harga", "Alias", "Catatan", "Emoji"]


def menus_to_workbook_bytes(menus: List[Dict[str, Any]]) -> bytes:
    """Build file Excel (bytes) dari list menu. Import openpyxl di sini
    biar tidak jadi hard dependency saat modul cuma di-import untuk
    read-only path."""

    import io
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Menu"

    ws.append(HEADERS)

    for i, m in enumerate(menus, 1):
        ws.append([
            m.get("no", i),
            m.get("resto") or "",
            m.get("category") or "",
            m.get("name") or "",
            m.get("price") or 0,
            m.get("alias") or "",
            m.get("note") or "",
            m.get("emoji") or "",
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def workbook_bytes_to_menus(data: bytes) -> List[Dict[str, Any]]:
    """Parse file Excel jadi list dict menu.

    Header case-insensitive; kolom bisa urut apa saja asal namanya
    dikenali. Field yang kosong dianggap string kosong / harga None."""

    import io
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    ws = wb["Menu"] if "Menu" in wb.sheetnames else wb.active

    if ws.max_row < 2:
        return []

    # Baca header, map ke indeks kolom
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header_map = {}

    aliases = {
        "no": "no",
        "resto": "resto",
        "kategori": "category",
        "category": "category",
        "nama": "name",
        "name": "name",
        "harga": "price",
        "price": "price",
        "alias": "alias",
        "catatan": "note",
        "note": "note",
        "emoji": "emoji",
    }

    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        key = str(cell).strip().lower()
        canonical = aliases.get(key)
        if canonical:
            header_map[canonical] = idx

    if "name" not in header_map or "price" not in header_map:
        raise ValueError(
            "File Excel harus punya kolom 'Nama' dan 'Harga'. "
            "Cek header baris pertama."
        )

    def val(row, key):
        idx = header_map.get(key)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    menus: List[Dict[str, Any]] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        # Skip baris kosong (semua kolom None/kosong)
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
            continue

        menus.append({
            "no": val(row, "no"),
            "resto": val(row, "resto"),
            "category": val(row, "category"),
            "name": val(row, "name"),
            "price": val(row, "price"),
            "alias": val(row, "alias"),
            "note": val(row, "note"),
            "emoji": val(row, "emoji"),
        })

    return menus

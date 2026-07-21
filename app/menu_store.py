"""
menu_store.py

Penyimpanan menu ke SQLite database (data/orders.db) supaya menu tidak hilang
saat redeploy di Render (yang punya ephemeral file system).

Migrasi dari Excel (data/menu.xlsx):
- Saat bot start pertama kali, import menu dari Excel ke DB otomatis
- Setelah itu semua perubahan menu disimpan ke DB (tidak ke Excel lagi)
- DB ter-persist dan di-backup otomatis seperti orders.db
"""

import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

try:
    from .constants import HOME_RESTO
except ImportError:
    try:
        from constants import HOME_RESTO
    except ImportError:
        HOME_RESTO = "MADAM LILY"


DEFAULT_DB_PATH = Path(__file__).parent.parent / "data" / "orders.db"


def _now_iso():
    """Fallback timestamp function (avoid circular import)."""
    return datetime.now().isoformat()


class MenuStore:
    """CRUD menu dengan backend SQLite database."""

    def __init__(self, db_path=None):
        self.db_path = Path(db_path) if db_path else DEFAULT_DB_PATH
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        self._init_schema()
        self._auto_migrate_from_excel()

    def _connect(self):
        conn = sqlite3.connect(self.db_path, check_same_thread=False)
        conn.row_factory = sqlite3.Row
        return conn

    def _init_schema(self):
        """Buat tabel menus di database jika belum ada."""
        with self._connect() as conn:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS menus (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    no INTEGER NOT NULL,
                    resto TEXT NOT NULL,
                    category TEXT DEFAULT '',
                    name TEXT UNIQUE NOT NULL,
                    price INTEGER NOT NULL,
                    alias TEXT DEFAULT '',
                    note TEXT DEFAULT '',
                    emoji TEXT DEFAULT '',
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL
                )
                """
            )
            conn.execute(
                "CREATE INDEX IF NOT EXISTS idx_menus_name ON menus(name)"
            )
            conn.execute(
                "CREATE INDEX IF NOT EXISTS idx_menus_resto ON menus(resto)"
            )
            conn.commit()

    def _auto_migrate_from_excel(self):
        """Cek: kalau DB masih kosong, import menu dari Excel ke DB."""
        try:
            with self._connect() as conn:
                count = conn.execute("SELECT COUNT(*) as c FROM menus").fetchone()["c"]

            if count == 0:
                # Try import dari Excel, tapi fail silent jika tidak ada
                excel_path = Path(__file__).parent.parent / "data" / "menu.xlsx"
                if excel_path.exists():
                    try:
                        import openpyxl
                        wb = openpyxl.load_workbook(excel_path, data_only=True)
                        ws = wb["Menu"] if "Menu" in wb.sheetnames else wb.active

                        with self._connect() as conn:
                            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
                                _, resto, kategori, nama, harga, alias, catatan, emoji = (
                                    list(row) + [None] * 8
                                )[:8]

                                if not nama or harga is None:
                                    continue

                                try:
                                    conn.execute(
                                        """
                                        INSERT INTO menus (no, resto, category, name, price, alias, note, emoji, created_at, updated_at)
                                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                                        """,
                                        (
                                            row_idx, resto or HOME_RESTO, kategori or "", nama,
                                            int(harga), alias or None, catatan or None, emoji or None,
                                            _now_iso(), _now_iso()
                                        ),
                                    )
                                except sqlite3.IntegrityError:
                                    pass  # Skip duplicate
                            conn.commit()
                    except Exception as e:
                        print(f"Warning: auto-migrate menu dari Excel gagal: {e}")
        except Exception as e:
            print(f"Warning: MenuStore initialization gagal: {e}")

    def add_menu(
        self,
        name: str,
        price: int,
        category: str = "",
        resto: str = "",
        emoji: str = "",
        alias: str = "",
        note: str = "",
    ) -> Tuple[bool, str]:
        """Tambah menu baru ke database."""
        if self.find_menu(name):
            return False, f"Menu '{name}' sudah ada di daftar."

        resto = resto or HOME_RESTO
        now = _now_iso()

        try:
            with self._connect() as conn:
                # Cari no terbesar di DB
                result = conn.execute(
                    "SELECT MAX(no) as max_no FROM menus"
                ).fetchone()
                next_no = (result["max_no"] or 0) + 1

                conn.execute(
                    """
                    INSERT INTO menus (no, resto, category, name, price, alias, note, emoji, created_at, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (next_no, resto, category, name, price, alias or None, note or None, emoji or None, now, now),
                )
                conn.commit()

            return True, (
                f"Menu '{name}' berhasil ditambahkan.\n"
                f"Harga: {price:,} Riel\n"
                f"Kategori: {category or '(kosong)'}\n"
                f"Resto: {resto}"
            )
        except Exception as e:
            return False, f"Error saat tambah menu: {e}"

    def remove_menu(self, name: str) -> Tuple[bool, str]:
        """Hapus menu dari database."""
        menu = self.find_menu(name)
        if not menu:
            return False, f"Menu '{name}' tidak ditemukan di daftar."

        try:
            with self._connect() as conn:
                conn.execute("DELETE FROM menus WHERE name = ?", (name,))
                conn.commit()

            return True, f"Menu '{name}' ({menu['price']:,} Riel) berhasil dihapus."
        except Exception as e:
            return False, f"Error saat hapus menu: {e}"

    def update_price(self, name: str, new_price: int) -> Tuple[bool, str]:
        """Update harga menu."""
        menu = self.find_menu(name)
        if not menu:
            return False, f"Menu '{name}' tidak ditemukan di daftar."

        old_price = menu["price"]
        now = _now_iso()

        try:
            with self._connect() as conn:
                conn.execute(
                    "UPDATE menus SET price = ?, updated_at = ? WHERE name = ?",
                    (new_price, now, name),
                )
                conn.commit()

            return True, (
                f"Harga '{name}' diperbarui:\n"
                f"{int(old_price):,} Riel -> {new_price:,} Riel"
            )
        except Exception as e:
            return False, f"Error saat update harga: {e}"

    def find_menu(self, name: str) -> Optional[Dict[str, Any]]:
        """Cari menu by name (case-insensitive)."""
        name_upper = name.strip().upper()

        with self._connect() as conn:
            result = conn.execute(
                "SELECT * FROM menus WHERE UPPER(name) = ?", (name_upper,)
            ).fetchone()

        if not result:
            return None

        return {
            "id": result["id"],
            "no": result["no"],
            "name": result["name"],
            "price": result["price"],
            "category": result["category"],
            "resto": result["resto"],
            "emoji": result["emoji"],
            "alias": result["alias"],
            "note": result["note"],
        }

    def get_all_menus(self) -> List[Dict[str, Any]]:
        """Return semua menu dalam format untuk parser/adapter."""
        with self._connect() as conn:
            rows = conn.execute("SELECT * FROM menus ORDER BY no ASC").fetchall()

        return [
            {
                "id": r["id"],
                "no": r["no"],
                "name": r["name"],
                "price": r["price"],
                "category": r["category"],
                "resto": r["resto"],
                "emoji": r["emoji"],
                "alias": r["alias"],
                "note": r["note"],
            }
            for r in rows
        ]

    def list_categories(self) -> List[Tuple[str, int]]:
        """Return daftar kategori dengan jumlah menu."""
        with self._connect() as conn:
            rows = conn.execute(
                "SELECT category, COUNT(*) as cnt FROM menus WHERE category != '' GROUP BY category ORDER BY category"
            ).fetchall()

        return [(r["category"], r["cnt"]) for r in rows]

    def list_menus_by_category(self, category: str) -> List[Dict[str, Any]]:
        """Return menu dalam kategori tertentu."""
        if not category:
            return []
        cat_upper = category.strip().upper()
        menus = self.get_all_menus()
        return [m for m in menus if cat_upper in (m.get("category") or "").upper()]

    def search_menus(self, keyword: str) -> List[Dict[str, Any]]:
        """Cari menu by keyword."""
        kw = keyword.strip().upper()
        menus = self.get_all_menus()
        return [m for m in menus if kw in m["name"].upper()]

    def bulk_upsert(
        self, rows: List[Dict[str, Any]]
    ) -> Tuple[int, int, List[str]]:
        """Insert/update banyak menu sekaligus (upsert by name).

        Nama menu (case-insensitive) jadi kunci: kalau sudah ada,
        update field-nya; kalau belum, insert baris baru. TIDAK
        menghapus menu lain -- pakai /hapusmenu untuk itu supaya salah
        upload file tidak menghilangkan semua menu.

        Return (jumlah_ditambah, jumlah_diupdate, errors)."""

        added = 0
        updated = 0
        errors: List[str] = []

        with self._connect() as conn:
            # No terbesar dipakai buat entry baru
            row = conn.execute("SELECT MAX(no) as m FROM menus").fetchone()
            next_no = (row["m"] or 0) + 1
            now = _now_iso()

            for i, r in enumerate(rows, 1):
                name = str(r.get("name") or "").strip()
                price = r.get("price")

                if not name:
                    errors.append(f"Baris {i}: nama kosong.")
                    continue

                if price is None:
                    errors.append(f"Baris {i} ({name}): harga kosong.")
                    continue

                try:
                    price = int(price)
                except (TypeError, ValueError):
                    errors.append(f"Baris {i} ({name}): harga bukan angka.")
                    continue

                resto = str(r.get("resto") or HOME_RESTO).strip() or HOME_RESTO
                category = str(r.get("category") or "").strip()
                alias = str(r.get("alias") or "").strip()
                note = str(r.get("note") or "").strip()
                emoji = str(r.get("emoji") or "").strip()

                existing = conn.execute(
                    "SELECT id FROM menus WHERE UPPER(name) = UPPER(?)",
                    (name,),
                ).fetchone()

                try:
                    if existing:
                        conn.execute(
                            """
                            UPDATE menus
                            SET resto = ?, category = ?, price = ?,
                                alias = ?, note = ?, emoji = ?, updated_at = ?
                            WHERE id = ?
                            """,
                            (
                                resto, category, price,
                                alias or None, note or None, emoji or None,
                                now, existing["id"],
                            ),
                        )
                        updated += 1
                    else:
                        conn.execute(
                            """
                            INSERT INTO menus
                            (no, resto, category, name, price, alias, note, emoji,
                             created_at, updated_at)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            """,
                            (
                                next_no, resto, category, name, price,
                                alias or None, note or None, emoji or None,
                                now, now,
                            ),
                        )
                        next_no += 1
                        added += 1
                except sqlite3.Error as e:
                    errors.append(f"Baris {i} ({name}): DB error {e}")

            conn.commit()

        return added, updated, errors
